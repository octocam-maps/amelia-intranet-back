"""
Excel del resumen mensual del técnico (requerimiento v1.2 §M1).

No se prueba "que el fichero exista": se prueba que los cuatro totales que
RRHH pidió estén, que las horas sean sumables y que el ×1,45 sea una fórmula
—no un número ya calculado— porque es el documento con el que se liquida
tiempo de descanso de una persona.
"""

import io
from datetime import date, datetime, timedelta, timezone

from openpyxl import load_workbook

from src.features.time_clock.application.results import TechnicianMonthSummary
from src.features.time_clock.domain.entities import (
    OvernightStay,
    ProductCategory,
    TechnicianDailyLog,
)
from src.features.time_clock.infrastructure.technician_xlsx_export import (
    build_technician_month_workbook,
    month_filename,
)

MADRID = timezone(timedelta(hours=2))


def _log(day: int, start_hour: int, end_hour: int, overnight: OvernightStay, *, crosses=False):
    work_date = date(2026, 7, day)
    end_day = day + 1 if crosses else day
    return TechnicianDailyLog(
        entry_id=f"e-{day}",
        user_id="u-1",
        work_date=work_date,
        started_at=datetime(2026, 7, day, start_hour, 0, tzinfo=MADRID),
        ended_at=datetime(2026, 7, end_day, end_hour, 0, tzinfo=MADRID),
        project_id="p-1",
        work_location="Guadix, Granada",
        had_break=False,
        break_minutes=0,
        overnight_stay=overnight,
        product_category=ProductCategory.HARDWARE,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc),
        project_name="Planta Guadix",
        full_name="Ana Ruiz",
    )


def _summary(**overrides) -> TechnicianMonthSummary:
    defaults = {
        "year": 2026,
        "month": 7,
        "budget_minutes": 9720,
        "worked_minutes": 10800,
        "overtime_minutes": 1080,
        "compensation_minutes": 1566,
        "overnight_stays_spain": 4,
        "overnight_stays_abroad": 2,
        "is_closed": True,
    }
    return TechnicianMonthSummary(**{**defaults, **overrides})


def _build(logs=None, summary=None):
    content = build_technician_month_workbook(
        logs if logs is not None else [_log(1, 8, 20, OvernightStay.ESPANA)],
        summary or _summary(),
        technician_name="Ana Ruiz",
    )
    return load_workbook(io.BytesIO(content))


def test_the_workbook_has_a_detail_sheet_and_a_summary_sheet():
    assert _build().sheetnames == ["Detalle", "Resumen"]


def test_hours_are_real_excel_durations_so_hr_can_sum_them():
    """Como texto no se pueden sumar y como decimal obligan a saber que ,75
    son 45 minutos. RRHH cuadra la nómina sumando esta columna a mano."""
    ws = _build()["Detalle"]
    hours_cell = ws.cell(row=6, column=9)

    assert isinstance(hours_cell.value, timedelta)
    assert hours_cell.value == timedelta(hours=12)
    assert hours_cell.number_format == "[h]:mm"


def test_the_1_45_factor_is_a_live_formula_not_a_precomputed_number():
    """Un documento de liquidación tiene que poder auditarse pinchando la
    celda. Si el número viniera hecho, comprobarlo exigiría rehacer la cuenta
    aparte."""
    ws = _build()["Resumen"]
    labels = {ws.cell(row=r, column=1).value: r for r in range(5, 20)}
    row = labels["Compensación (horas extra × 1,45)"]

    assert ws.cell(row=row, column=2).value == "=B7*1.45"
    # Y B7 debe ser efectivamente "Total horas extra": una fórmula que apunte
    # a la celda equivocada es peor que no tener fórmula.
    assert ws.cell(row=7, column=1).value == "Total horas extra"


def test_the_factor_label_uses_a_spanish_decimal_comma():
    ws = _build()["Resumen"]
    labels = [ws.cell(row=r, column=1).value for r in range(5, 20)]

    assert "Compensación (horas extra × 1,45)" in labels
    assert "Compensación (horas extra × 1.45)" not in labels


def test_the_summary_reports_the_four_totals_hr_asked_for():
    ws = _build()["Resumen"]
    values = {
        ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
        for r in range(5, 20)
        if ws.cell(row=r, column=1).value
    }

    assert values["Total horas extra"] == timedelta(minutes=1080)
    assert values["Total pernoctas en España"] == 4
    assert values["Total pernoctas fuera de España"] == 2
    assert values["Total pernoctas"] == 6


def test_overnight_stay_is_shown_as_two_columns_even_though_it_is_one_field():
    """RRHH lo pidió como dos preguntas; el modelo lo guarda como un enum. El
    Excel vuelve a presentarlo como dos columnas."""
    logs = [
        _log(1, 8, 20, OvernightStay.ESPANA),
        _log(2, 8, 20, OvernightStay.EXTRANJERO),
        _log(3, 8, 20, OvernightStay.NINGUNA),
    ]
    ws = _build(logs)["Detalle"]

    assert [ws.cell(row=r, column=10).value for r in (6, 7, 8)] == ["Sí", "Sí", "No"]
    assert [ws.cell(row=r, column=11).value for r in (6, 7, 8)] == [
        "España",
        "Fuera de España",
        "—",
    ]


def test_a_shift_crossing_midnight_keeps_its_start_date_and_full_duration():
    logs = [_log(5, 8, 1, OvernightStay.ESPANA, crosses=True)]
    ws = _build(logs)["Detalle"]

    assert ws.cell(row=6, column=1).value == "05/07/2026"
    assert ws.cell(row=6, column=5).value == "08:00"
    assert ws.cell(row=6, column=6).value == "01:00"
    assert ws.cell(row=6, column=9).value == timedelta(hours=17)


def test_an_open_month_warns_that_overtime_may_still_change():
    ws = _build(summary=_summary(is_closed=False))["Resumen"]
    texts = [ws.cell(row=r, column=1).value for r in range(5, 22)]

    assert any(t and "todavía no ha terminado" in t for t in texts)


def test_the_filename_carries_technician_and_month():
    assert month_filename(_summary(), "Ana Ruiz") == "registro-horario-ana-ruiz-2026-07.xlsx"

"""No parsea el `.xlsx` byte a byte — solo comprueba que `openpyxl` puede
reabrir lo que generamos y que el contenido de negocio (cabecera, filas,
heurística Nombre/Apellido, logo) llega donde se espera."""

import io
from datetime import date, datetime, timezone

from openpyxl import load_workbook

from src.features.time_clock.domain.entities import TimeClockExportRow
from src.features.time_clock.infrastructure.xlsx_export import (
    TITLE_ADMIN,
    TITLE_EMPLOYEE,
    _split_full_name,
    build_time_clock_export_workbook,
)


def _row(
    full_name: str, with_clock_out: bool = True, source: str = "manual"
) -> TimeClockExportRow:
    clock_in = datetime(2026, 7, 9, 8, 0, tzinfo=timezone.utc)
    clock_out = datetime(2026, 7, 9, 16, 30, tzinfo=timezone.utc) if with_clock_out else None
    return TimeClockExportRow(
        user_id="user-1",
        full_name=full_name,
        dni_nif="12345678A",
        phone="600111222",
        work_date=date(2026, 7, 9),
        clock_in=clock_in,
        clock_out=clock_out,
        source=source,
    )


def test_split_full_name_first_word_is_nombre_rest_is_apellido():
    assert _split_full_name("Ana García López") == ("Ana", "García López")
    assert _split_full_name("Beatriz") == ("Beatriz", "")


def test_workbook_has_header_and_data_rows():
    rows = [
        _row("Ana García", source="manual"),
        _row("Luis Pérez", with_clock_out=False, source="live"),
    ]

    workbook_bytes = build_time_clock_export_workbook(
        rows, date_from=date(2026, 6, 15), date_to=date(2026, 7, 15)
    )

    wb = load_workbook(io.BytesIO(workbook_bytes))
    ws = wb.active

    header_values = [cell.value for cell in ws[5]]
    assert header_values == [
        "Nombre",
        "Apellido",
        "DNI",
        "Teléfono",
        "Fecha",
        "Entrada",
        "Salida",
        "Horas trabajadas",
        "Origen",
    ]

    first_data_row = [cell.value for cell in ws[6]]
    assert first_data_row[0] == "Ana"
    assert first_data_row[1] == "García"
    assert first_data_row[2] == "12345678A"
    # 08:00/16:30 UTC -> 10:00/18:30 en Madrid (CEST, verano) — se muestra en
    # hora local de España, no en la hora UTC almacenada.
    assert first_data_row[5] == "10:00"
    assert first_data_row[6] == "18:30"
    assert first_data_row[7] == 8.5  # duración absoluta: no depende de la TZ de visualización
    # LOGIC-2 (pentest ético): RRHH necesita distinguir autodeclarado (alta
    # manual) de fichado en vivo directamente en el informe exportable.
    assert first_data_row[8] == "Manual"

    open_entry_row = [cell.value for cell in ws[7]]
    assert not open_entry_row[6]  # sin salida (openpyxl relee "" como None)
    assert open_entry_row[7] == 0.0  # tramo en curso no suma horas al informe
    assert open_entry_row[8] == "En vivo"


def test_workbook_title_defaults_to_admin_scope():
    workbook_bytes = build_time_clock_export_workbook(
        [_row("Ana García")], date_from=date(2026, 6, 15), date_to=date(2026, 7, 15)
    )

    wb = load_workbook(io.BytesIO(workbook_bytes))
    assert wb.active["A2"].value == TITLE_ADMIN


def test_workbook_title_uses_employee_scope_when_passed():
    workbook_bytes = build_time_clock_export_workbook(
        [_row("Ana García")],
        date_from=date(2026, 6, 15),
        date_to=date(2026, 7, 15),
        title=TITLE_EMPLOYEE,
    )

    wb = load_workbook(io.BytesIO(workbook_bytes))
    assert wb.active["A2"].value == TITLE_EMPLOYEE


def test_workbook_has_logo_image_and_frozen_header():
    workbook_bytes = build_time_clock_export_workbook(
        [_row("Ana García")], date_from=date(2026, 6, 15), date_to=date(2026, 7, 15)
    )

    wb = load_workbook(io.BytesIO(workbook_bytes))
    ws = wb.active

    assert len(ws._images) == 1
    assert ws.freeze_panes == "A6"

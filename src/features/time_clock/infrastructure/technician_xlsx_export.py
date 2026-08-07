"""
Resumen mensual del técnico en `.xlsx` (requerimiento v1.2 §M1: «a final de
cada mes se podrá descargar un Excel con el resumen de todos los registros,
indicando aparte del detalle completo... el total de pernoctas en España y el
total de pernoctas fuera de España, más total de horas extra, más resultado
del total de las horas extra por 1,45»).

Reutiliza el andamiaje de marca de `xlsx_export.py` (logo, cabecera navy,
título) en vez de duplicarlo: si RRHH cambia el logo, cambia en los dos
informes a la vez.

DOS DECISIONES DE FORMATO QUE NO SON COSMÉTICAS:

1. Las horas van como DURACIÓN REAL de Excel (`[h]:mm`), no como texto ni como
   decimal. RRHH suma columnas a mano en la propia hoja para cuadrar la
   nómina; "11h 45m" en texto no se puede sumar, y 11,75 obliga a saber que
   ese ,75 son 45 minutos.
2. El `× 1,45` se escribe como FÓRMULA VIVA que referencia la celda de horas
   extra, no como un número ya calculado. Un documento con el que se liquida
   tiempo de descanso tiene que poder auditarse pinchando la celda: si el
   número viniera hecho, comprobarlo exigiría rehacer la cuenta aparte.
"""

import io
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.shared.utils.timezone import MADRID_TZ

from ..application.results import TechnicianMonthSummary
from ..domain.entities import OvernightStay, ProductCategory, TechnicianDailyLog
from ..domain.policy import OVERTIME_COMPENSATION_FACTOR
from .xlsx_export import (
    _BRAND_NAVY,
    _HEADER_TEXT_COLOR,
    _SUBTITLE_TEXT_COLOR,
    _insert_logo,
)

_DETAIL_COLUMNS = [
    ("Fecha", 12),
    ("Técnico", 22),
    ("Proyecto", 22),
    ("Lugar de trabajo", 24),
    ("Inicio", 10),
    ("Fin", 10),
    ("¿Pausa?", 10),
    ("Min. pausa", 12),
    ("Horas efectivas", 16),
    ("¿Pernocta?", 12),
    ("Lugar pernocta", 16),
    ("Producto", 12),
]

_HOURS_FORMAT = "[h]:mm"

_OVERNIGHT_LABELS = {
    OvernightStay.NINGUNA: ("No", "—"),
    OvernightStay.ESPANA: ("Sí", "España"),
    OvernightStay.EXTRANJERO: ("Sí", "Fuera de España"),
}

_PRODUCT_LABELS = {
    ProductCategory.SOFTWARE: "Software",
    ProductCategory.HARDWARE: "Hardware",
}

_MONTH_NAMES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]

_LOGO_ROW = 1
_TITLE_ROW = 2
_SUBTITLE_ROW = 3
_HEADER_ROW = 5
_DATA_START_ROW = 6


def _as_duration(minutes: int) -> timedelta:
    """openpyxl serializa un `timedelta` como duración nativa de Excel, que es
    lo que hace la columna sumable con `[h]:mm`."""
    return timedelta(minutes=minutes)


def _local_time(value: datetime) -> str:
    """Hora de pared en Madrid. Los `TIMESTAMPTZ` llegan en UTC y mostrarlos
    tal cual pondría "06:00" donde la persona fichó a las 08:00."""
    return value.astimezone(MADRID_TZ).strftime("%H:%M")


def _write_title(ws: Worksheet, title: str, subtitle: str, last_column: int) -> None:
    ws.merge_cells(
        start_row=_TITLE_ROW, start_column=1, end_row=_TITLE_ROW, end_column=last_column
    )
    cell = ws.cell(row=_TITLE_ROW, column=1, value=title)
    cell.font = Font(name="Calibri", size=14, bold=True, color=_BRAND_NAVY)

    ws.merge_cells(
        start_row=_SUBTITLE_ROW, start_column=1, end_row=_SUBTITLE_ROW, end_column=last_column
    )
    subtitle_cell = ws.cell(row=_SUBTITLE_ROW, column=1, value=subtitle)
    subtitle_cell.font = Font(name="Calibri", size=10, italic=True, color=_SUBTITLE_TEXT_COLOR)


def _write_detail_sheet(ws: Worksheet, logs: list[TechnicianDailyLog], subtitle: str) -> None:
    _insert_logo(ws)
    _write_title(ws, "Registro horario — detalle del mes", subtitle, len(_DETAIL_COLUMNS))

    fill = PatternFill(start_color=_BRAND_NAVY, end_color=_BRAND_NAVY, fill_type="solid")
    for col_index, (title, width) in enumerate(_DETAIL_COLUMNS, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=col_index, value=title)
        cell.font = Font(name="Calibri", bold=True, color=_HEADER_TEXT_COLOR)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_index)].width = width
    ws.row_dimensions[_HEADER_ROW].height = 22

    for offset, log in enumerate(logs):
        row = _DATA_START_ROW + offset
        had_overnight, overnight_place = _OVERNIGHT_LABELS[log.overnight_stay]
        values = [
            log.work_date.strftime("%d/%m/%Y"),
            log.full_name or "",
            log.project_name or "",
            log.work_location,
            _local_time(log.started_at),
            _local_time(log.ended_at),
            "Sí" if log.had_break else "No",
            log.break_minutes,
            _as_duration(log.worked_minutes),
            had_overnight,
            overnight_place,
            _PRODUCT_LABELS[log.product_category],
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_index, value=value)
            cell.alignment = Alignment(horizontal="left" if col_index in (2, 3, 4) else "center")
            if col_index == 9:
                cell.number_format = _HOURS_FORMAT

    ws.freeze_panes = f"A{_DATA_START_ROW}"


def _write_summary_sheet(
    ws: Worksheet, summary: TechnicianMonthSummary, subtitle: str
) -> None:
    _insert_logo(ws)
    _write_title(ws, "Registro horario — resumen del mes", subtitle, 2)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    label_font = Font(name="Calibri", bold=False)
    total_font = Font(name="Calibri", bold=True)

    rows: list[tuple[str, object, str | None, bool]] = [
        ("Total horas trabajadas en el mes", _as_duration(summary.worked_minutes),
         _HOURS_FORMAT, False),
        ("Bolsa mensual", _as_duration(summary.budget_minutes), _HOURS_FORMAT, False),
        ("Total horas extra", _as_duration(summary.overtime_minutes), _HOURS_FORMAT, True),
    ]

    current = _HEADER_ROW
    for label, value, number_format, is_total in rows:
        ws.cell(row=current, column=1, value=label).font = total_font if is_total else label_font
        cell = ws.cell(row=current, column=2, value=value)
        cell.font = total_font if is_total else label_font
        if number_format:
            cell.number_format = number_format
        current += 1

    # La fórmula viva: referencia la celda de horas extra, que es la fila
    # anterior. Si RRHH ajusta el detalle, este número se recalcula solo.
    #
    # La ETIQUETA lleva coma decimal (español) y la FÓRMULA lleva punto: el
    # separador de la fórmula es parte de la sintaxis del fichero xlsx, no del
    # idioma, y Excel ya la muestra localizada al abrirla.
    overtime_cell = f"B{current - 1}"
    factor_label = str(OVERTIME_COMPENSATION_FACTOR).replace(".", ",")
    ws.cell(
        row=current, column=1, value=f"Compensación (horas extra × {factor_label})"
    ).font = total_font
    compensation = ws.cell(
        row=current, column=2, value=f"={overtime_cell}*{OVERTIME_COMPENSATION_FACTOR}"
    )
    compensation.font = total_font
    compensation.number_format = _HOURS_FORMAT
    current += 2

    for label, value in (
        ("Total pernoctas en España", summary.overnight_stays_spain),
        ("Total pernoctas fuera de España", summary.overnight_stays_abroad),
        ("Total pernoctas", summary.overnight_stays_total),
    ):
        ws.cell(row=current, column=1, value=label).font = label_font
        ws.cell(row=current, column=2, value=value).font = label_font
        current += 1

    if not summary.is_closed:
        current += 1
        note = ws.cell(
            row=current,
            column=1,
            value="El mes todavía no ha terminado: las horas extra pueden variar.",
        )
        note.font = Font(name="Calibri", size=10, italic=True, color=_SUBTITLE_TEXT_COLOR)


def build_technician_month_workbook(
    logs: list[TechnicianDailyLog],
    summary: TechnicianMonthSummary,
    *,
    technician_name: str,
) -> bytes:
    month_label = f"{_MONTH_NAMES[summary.month - 1]} de {summary.year}"
    subtitle = f"{technician_name} — {month_label}"

    wb = Workbook()
    detail = wb.active
    detail.title = "Detalle"
    _write_detail_sheet(detail, logs, subtitle)

    resumen = wb.create_sheet("Resumen")
    _write_summary_sheet(resumen, summary, subtitle)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def month_filename(summary: TechnicianMonthSummary, technician_name: str) -> str:
    safe_name = "".join(c if c.isalnum() else "-" for c in technician_name).strip("-").lower()
    return f"registro-horario-{safe_name}-{summary.year}-{summary.month:02d}.xlsx"

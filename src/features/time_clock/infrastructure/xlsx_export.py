"""
Construye el libro `.xlsx` del informe de fichajes admin
(`GET /time-clock/entries/export.xlsx`). `openpyxl` vive SOLO aquí — es un
detalle de formato de salida (infraestructura), no lógica de negocio; el
caso de uso (`application/use_cases/export_time_clock_entries.py`) no sabe
que el resultado termina en un XLSX.
"""

import io
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.shared.utils.timezone import MADRID_TZ

from ..domain.entities import TimeClockExportRow

_BRAND_NAVY = "0F1729"
_HEADER_TEXT_COLOR = "FFFFFF"
_SUBTITLE_TEXT_COLOR = "6B7280"

# .../src/features/time_clock/infrastructure/xlsx_export.py -> .../src
_SRC_DIR = Path(__file__).resolve().parent.parent.parent.parent
_LOGO_PATH = _SRC_DIR / "shared" / "assets" / "brand" / "logo-amelia.png"

_COLUMN_TITLES = [
    "Nombre",
    "Apellido",
    "DNI",
    "Teléfono",
    "Fecha",
    "Entrada",
    "Salida",
    "Horas trabajadas",
]
_COLUMN_WIDTHS = [16, 20, 14, 16, 12, 10, 10, 18]

_LOGO_ROW = 1
_TITLE_ROW = 2
_SUBTITLE_ROW = 3
_HEADER_ROW = 5
_DATA_START_ROW = 6


def _split_full_name(full_name: str) -> tuple[str, str]:
    """Nombre = primera palabra; Apellido = el resto. `users.full_name` es un
    único campo de texto (no hay columnas Nombre/Apellido separadas en el
    esquema) — misma heurística que el `ORDER BY` del repositorio
    (`infrastructure/repositories/time_clock_repository.py::_EXPORT_SELECT`)
    para que el orden de las filas coincida con lo que se lee en pantalla."""
    parts = full_name.strip().split(" ", 1)
    nombre = parts[0] if parts else ""
    apellido = parts[1] if len(parts) > 1 else ""
    return nombre, apellido


def _fmt_time(value) -> str:
    """`HH:MM` en hora de Madrid — asyncpg devuelve `TIMESTAMPTZ` en UTC;
    RRHH opera en horario español (mismo criterio que `today_in_madrid()`)."""
    if value is None:
        return ""
    return value.astimezone(MADRID_TZ).strftime("%H:%M")


def _worked_hours(row: TimeClockExportRow) -> float:
    """Decimal (no `HH:MM`) para que Excel pueda sumar/promediar la columna
    directamente. Un tramo en curso (sin `clock_out`) exporta `0` — RRHH no
    puede cerrar un informe con jornadas todavía abiertas."""
    minutes = row.worked_minutes
    return round(minutes / 60, 2) if minutes is not None else 0.0


def _insert_logo(ws: Worksheet) -> None:
    # El PNG blanco (`logo-amelia-blanco.png`) no se ve sobre el fondo claro
    # de la hoja — se usa la versión de color (`logo-amelia.png`, la misma
    # que el frontend renderiza en fondos claros).
    if not _LOGO_PATH.exists():
        return
    image = XLImage(str(_LOGO_PATH))
    # Proporción original 1920x400 (4.8:1) — se escala a un tamaño de
    # cabecera razonable sin deformarla.
    image.width = 192
    image.height = 40
    ws.row_dimensions[_LOGO_ROW].height = 32
    ws.add_image(image, "A1")


def _write_title(ws: Worksheet, title: str, date_from: date, date_to: date) -> None:
    last_column = len(_COLUMN_TITLES)

    ws.merge_cells(start_row=_TITLE_ROW, start_column=1, end_row=_TITLE_ROW, end_column=last_column)
    title_cell = ws.cell(row=_TITLE_ROW, column=1, value=title)
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=_BRAND_NAVY)

    ws.merge_cells(
        start_row=_SUBTITLE_ROW, start_column=1, end_row=_SUBTITLE_ROW, end_column=last_column
    )
    subtitle_cell = ws.cell(
        row=_SUBTITLE_ROW,
        column=1,
        value=f"Del {date_from.strftime('%d/%m/%Y')} al {date_to.strftime('%d/%m/%Y')}",
    )
    subtitle_cell.font = Font(name="Calibri", size=10, italic=True, color=_SUBTITLE_TEXT_COLOR)


def _write_header(ws: Worksheet) -> None:
    fill = PatternFill(start_color=_BRAND_NAVY, end_color=_BRAND_NAVY, fill_type="solid")
    for col_index, title in enumerate(_COLUMN_TITLES, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=col_index, value=title)
        cell.font = Font(name="Calibri", bold=True, color=_HEADER_TEXT_COLOR)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[_HEADER_ROW].height = 22


def _write_rows(ws: Worksheet, rows: list[TimeClockExportRow]) -> None:
    for offset, row in enumerate(rows):
        excel_row = _DATA_START_ROW + offset
        nombre, apellido = _split_full_name(row.full_name)
        values = [
            nombre,
            apellido,
            row.dni_nif or "",
            row.phone or "",
            row.work_date.strftime("%d/%m/%Y"),
            _fmt_time(row.clock_in),
            _fmt_time(row.clock_out),
            _worked_hours(row),
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_index, value=value)
            cell.alignment = Alignment(horizontal="center" if col_index >= 5 else "left")
            if col_index == len(values):
                cell.number_format = "0.00"
        ws.row_dimensions[excel_row].height = 18


def _apply_column_widths(ws: Worksheet) -> None:
    for col_index, width in enumerate(_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width


# Públicos (sin `_`): el router (`infrastructure/routes.py`) los importa
# para decidir qué rótulo pasar según el rol del requester.
TITLE_ADMIN = "Registro de fichajes — toda la plantilla (últimos 30 días)"
TITLE_EMPLOYEE = "Mis fichajes — últimos 30 días"


def build_time_clock_export_workbook(
    rows: list[TimeClockExportRow],
    *,
    date_from: date,
    date_to: date,
    title: str = TITLE_ADMIN,
) -> bytes:
    """Genera el `.xlsx` completo en memoria (nunca toca disco salvo para
    leer el PNG estático del logo). `title` distingue el alcance del informe
    (admin: toda la plantilla / empleado: solo lo propio) — el resto del
    formato (logo, columnas, cabecera navy) es idéntico para ambos roles."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Fichajes"

    _insert_logo(ws)
    _write_title(ws, title, date_from, date_to)
    _write_header(ws)
    _write_rows(ws, rows)
    _apply_column_widths(ws)
    # Congela cabecera (+ logo/título por encima) al hacer scroll por las
    # ~850 filas de fichajes.
    ws.freeze_panes = f"A{_DATA_START_ROW}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

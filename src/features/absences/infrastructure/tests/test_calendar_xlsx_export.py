"""No parsea el `.xlsx` byte a byte — solo comprueba que `openpyxl` puede
reabrir lo que generamos y que el contenido de negocio (cabecera, filas,
logo) llega donde se espera. Mismo patrón que
`time_clock/infrastructure/tests/test_xlsx_export.py`."""

import io
from datetime import date

from openpyxl import load_workbook

from src.features.absences.domain.entities import AbsenceCalendarEntry
from src.features.absences.infrastructure.calendar_xlsx_export import (
    TITLE,
    build_absence_calendar_export_workbook,
)


def _entry(**overrides) -> AbsenceCalendarEntry:
    kwargs = dict(
        request_id="req-1",
        user_id="user-1",
        user_full_name="Ana García",
        absence_type_id="type-vacaciones",
        absence_type_name="Vacaciones",
        absence_type_color="#00D170",
        start_date=date(2026, 7, 20),
        end_date=date(2026, 7, 24),
        days_count=5.0,
        status="approved",
    )
    kwargs.update(overrides)
    return AbsenceCalendarEntry(**kwargs)


def test_workbook_has_header_and_data_rows():
    entries = [_entry(), _entry(user_full_name="Luis Pérez", status="pending", days_count=1.5)]

    workbook_bytes = build_absence_calendar_export_workbook(
        entries, date_from=date(2026, 7, 1), date_to=date(2026, 7, 31)
    )

    wb = load_workbook(io.BytesIO(workbook_bytes))
    ws = wb.active

    header_values = [cell.value for cell in ws[5]]
    assert header_values == ["Empleado", "Tipo de ausencia", "Inicio", "Fin", "Días", "Estado"]

    first_data_row = [cell.value for cell in ws[6]]
    assert first_data_row == ["Ana García", "Vacaciones", "20/07/2026", "24/07/2026", 5.0, "Aprobada"]

    second_data_row = [cell.value for cell in ws[7]]
    assert second_data_row[0] == "Luis Pérez"
    assert second_data_row[5] == "Pendiente"


def test_workbook_title_and_subtitle():
    workbook_bytes = build_absence_calendar_export_workbook(
        [_entry()], date_from=date(2026, 7, 1), date_to=date(2026, 7, 31)
    )

    wb = load_workbook(io.BytesIO(workbook_bytes))
    ws = wb.active
    assert ws["A2"].value == TITLE
    assert ws["A3"].value == "Del 01/07/2026 al 31/07/2026"


def test_workbook_has_logo_image_and_frozen_header():
    workbook_bytes = build_absence_calendar_export_workbook(
        [_entry()], date_from=date(2026, 7, 1), date_to=date(2026, 7, 31)
    )

    wb = load_workbook(io.BytesIO(workbook_bytes))
    ws = wb.active

    assert len(ws._images) == 1
    assert ws.freeze_panes == "A6"


def test_workbook_with_no_entries_only_has_header():
    workbook_bytes = build_absence_calendar_export_workbook(
        [], date_from=date(2026, 7, 1), date_to=date(2026, 7, 31)
    )

    wb = load_workbook(io.BytesIO(workbook_bytes))
    ws = wb.active
    assert ws["A6"].value is None


def test_workbook_without_subject_name_is_byte_identical_to_export_global():
    """RF-A1: `subject_name=None` (export global) DEBE mantener el output
    actual byte a byte — este es el punto donde es fácil romper algo que ya
    funcionaba."""
    kwargs = dict(date_from=date(2026, 7, 1), date_to=date(2026, 7, 31))

    without_kwarg = build_absence_calendar_export_workbook([_entry()], **kwargs)
    with_explicit_none = build_absence_calendar_export_workbook(
        [_entry()], subject_name=None, **kwargs
    )

    assert without_kwarg == with_explicit_none

    wb = load_workbook(io.BytesIO(without_kwarg))
    ws = wb.active
    assert ws["A4"].value is None


def test_workbook_with_subject_name_writes_employee_line():
    workbook_bytes = build_absence_calendar_export_workbook(
        [_entry()],
        date_from=date(2026, 7, 1),
        date_to=date(2026, 7, 31),
        subject_name="Ana García",
    )

    wb = load_workbook(io.BytesIO(workbook_bytes))
    ws = wb.active
    assert ws["A4"].value == "Empleado: Ana García"
    # El resto del layout no se desplaza: cabecera y datos siguen en las
    # mismas filas que el export global.
    assert ws["A5"].value == "Empleado"
    assert ws["A6"].value == "Ana García"

"""
Construye el libro `.xlsx` del "Calendario general de la plantilla"
(`GET /absences/calendar/export.xlsx`, LOTE 4). `openpyxl` vive SOLO aquí —
mismo patrón que `time_clock/infrastructure/xlsx_export.py` (logo, cabecera
navy, panel congelado): es un detalle de formato de salida (infraestructura),
el caso de uso (`GetAbsenceCalendarUseCase`) no sabe que el resultado termina
en un XLSX.
"""

import io
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..domain.entities import AbsenceCalendarEntry

_BRAND_NAVY = "0F1729"
_HEADER_TEXT_COLOR = "FFFFFF"
_SUBTITLE_TEXT_COLOR = "6B7280"

# .../src/features/absences/infrastructure/calendar_xlsx_export.py -> .../src
_SRC_DIR = Path(__file__).resolve().parent.parent.parent.parent
_LOGO_PATH = _SRC_DIR / "shared" / "assets" / "brand" / "logo-amelia.png"

_COLUMN_TITLES = ["Empleado", "Tipo de ausencia", "Inicio", "Fin", "Días", "Estado"]
_COLUMN_WIDTHS = [26, 22, 14, 14, 10, 14]

_LOGO_ROW = 1
_TITLE_ROW = 2
_SUBTITLE_ROW = 3
_HEADER_ROW = 5
_DATA_START_ROW = 6

_STATUS_LABELS = {"pending": "Pendiente", "approved": "Aprobada"}

TITLE = "Calendario de ausencias — toda la plantilla"


def _status_label(status: str) -> str:
    return _STATUS_LABELS.get(status, status.capitalize())


def _insert_logo(ws: Worksheet) -> None:
    if not _LOGO_PATH.exists():
        return
    image = XLImage(str(_LOGO_PATH))
    # Proporción original 1920x400 (4.8:1) — mismo tamaño de cabecera que el
    # export de fichajes (`xlsx_export.py`), para que ambos informes se vean
    # consistentes.
    image.width = 192
    image.height = 40
    ws.row_dimensions[_LOGO_ROW].height = 32
    ws.add_image(image, "A1")


def _write_title(ws: Worksheet, date_from: date, date_to: date) -> None:
    last_column = len(_COLUMN_TITLES)

    ws.merge_cells(start_row=_TITLE_ROW, start_column=1, end_row=_TITLE_ROW, end_column=last_column)
    title_cell = ws.cell(row=_TITLE_ROW, column=1, value=TITLE)
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=_BRAND_NAVY)

    ws.merge_cells(
        start_row=_SUBTITLE_ROW, start_column=1, end_row=_SUBTITLE_ROW, end_column=last_column
    )
    subtitle_cell = ws.cell(
        row=_SUBTITLE_ROW,
        column=1,
        value=f"Del {date_from.strftime('%d/%m/%Y')} al {date_to.strftime('%d/%m/%Y')}",
    )
    subtitle_cell.font = Font(name="Calibri", size=10, italic=True, color=_SUBTITLE_TEXT_COLOR)


def _write_header(ws: Worksheet) -> None:
    fill = PatternFill(start_color=_BRAND_NAVY, end_color=_BRAND_NAVY, fill_type="solid")
    for col_index, title in enumerate(_COLUMN_TITLES, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=col_index, value=title)
        cell.font = Font(name="Calibri", bold=True, color=_HEADER_TEXT_COLOR)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[_HEADER_ROW].height = 22


def _write_rows(ws: Worksheet, entries: list[AbsenceCalendarEntry]) -> None:
    for offset, entry in enumerate(entries):
        excel_row = _DATA_START_ROW + offset
        values = [
            entry.user_full_name,
            entry.absence_type_name,
            entry.start_date.strftime("%d/%m/%Y"),
            entry.end_date.strftime("%d/%m/%Y"),
            entry.days_count,
            _status_label(entry.status),
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_index, value=value)
            cell.alignment = Alignment(horizontal="center" if col_index >= 3 else "left")
            if col_index == 5:
                cell.number_format = "0.##"
        ws.row_dimensions[excel_row].height = 18


def _apply_column_widths(ws: Worksheet) -> None:
    for col_index, width in enumerate(_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width


def build_absence_calendar_export_workbook(
    entries: list[AbsenceCalendarEntry], *, date_from: date, date_to: date
) -> bytes:
    """Genera el `.xlsx` completo en memoria (nunca toca disco salvo para
    leer el PNG estático del logo). Mismo alcance que la pantalla y el PDF
    (`calendar_pdf_export.py`): TODA la plantilla, `pending`/`approved`,
    dentro del rango `[date_from, date_to]`."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendario"

    _insert_logo(ws)
    _write_title(ws, date_from, date_to)
    _write_header(ws)
    _write_rows(ws, entries)
    _apply_column_widths(ws)
    # Congela cabecera (+ logo/título por encima) al hacer scroll por la
    # plantilla completa.
    ws.freeze_panes = f"A{_DATA_START_ROW}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
